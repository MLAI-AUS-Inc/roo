#!/usr/bin/env python3
"""
luma_stripe_merge.py
====================
Pull Luma ticket sales + Stripe payments, merge them, and write a Xero
reconciliation pack:

  1. an .xlsx worksheet  (Sales detail + Payout summary)  — the audit trail
  2. a Cowork brief (.md) — one section per Stripe payout = one bank deposit,
     pre-filled with the Xero "Create" fields (Who / What / Why / Event Name /
     Project Name / Tax Rate) so Claude Cowork can reconcile it in the browser.

Because Luma processes card payments through your own Stripe account, the same
sale appears in both systems. This script joins them so each ticket sale is
linked to its Stripe charge and rolled up into the Stripe *payout* that lands in
your bank — which is the line you actually reconcile in Xero.

The pull is *payout-driven*: it starts from every payout that settled in the
window, pulls every charge inside it, then enriches with Luma. That guarantees
each payout's lines sum to the exact bank deposit (even if a charge's Luma sale
happened before the window).

Usage
-----
  # See the output shape with realistic fake data (no keys needed):
  python luma_stripe_merge.py --mock --out sample.xlsx

  # Real run (reads keys from environment / .env):
  python luma_stripe_merge.py --days 30 --out last30.xlsx
  python luma_stripe_merge.py --since 2026-06-01 --until 2026-06-30 --out june.xlsx
  python luma_stripe_merge.py --month 2026-06 --out june.xlsx

Auth (real run)
---------------
  LUMA_API_KEY     Luma calendar/org API key (requires Luma Plus)
  STRIPE_API_KEY   Stripe restricted key, read-only: Charges, Balance
                   transactions, Payouts, Customers
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
from collections import defaultdict

LUMA_BASE = "https://public-api.luma.com/v1"
STRIPE_BASE = "https://api.stripe.com/v1"

# --- Xero mapping (edit to match your chart of accounts) ----------------------
# These feed the "What (account)" column of the Cowork brief.
TICKET_INCOME_ACCOUNT = "Ticket Sales"   # income account ticket revenue posts to
STRIPE_FEE_ACCOUNT = "Stripe Fees"       # expense account for Stripe processing fees
LUMA_FEE_ACCOUNT = "Luma Fees"           # expense account for Luma's platform fee
DEFAULT_PAYER = "Stripe Payments"        # "Who" when a payout batches many buyers
TAX_RATE_LABEL = "account default"       # left as the income account's default; confirm in Xero
# Luma processes charges before the window can settle as a payout. Pull Luma
# sales this many days before the window start so recent payouts are fully named.
LUMA_LOOKBACK_DAYS = 45


def _money(x) -> float:
    return round(float(x or 0), 2)


def _load_dotenv() -> None:
    """Minimal .env loader so you don't need python-dotenv."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    for line in open(path):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def _load_event_map() -> dict:
    """Optional Luma-event-name -> Xero-Event-Name map (event_map.json)."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "event_map.json")
    if os.path.exists(path):
        try:
            return json.load(open(path))
        except (ValueError, OSError):
            return {}
    return {}


def _month_bounds(month: str) -> tuple[int, int]:
    """'YYYY-MM' -> (start_unix, end_unix) covering that calendar month (UTC)."""
    y, m = (int(p) for p in month.split("-"))
    start = dt.datetime(y, m, 1, tzinfo=dt.timezone.utc)
    end = dt.datetime(y + (m == 12), (m % 12) + 1, 1, tzinfo=dt.timezone.utc)
    return int(start.timestamp()), int(end.timestamp())


def _parse_date(s: str | None):
    if not s:
        return None
    return dt.datetime.fromisoformat(s).replace(tzinfo=dt.timezone.utc)


def resolve_window(args) -> tuple[int, int]:
    """Turn --month / --since/--until / --days into a (start, end) UTC unix pair."""
    if args.month:
        return _month_bounds(args.month)
    now = dt.datetime.now(dt.timezone.utc)
    if args.since or args.until:
        start = _parse_date(args.since) or (now - dt.timedelta(days=30))
        end = _parse_date(args.until) or now
    else:
        end = now
        start = now - dt.timedelta(days=args.days or 30)
    return int(start.timestamp()), int(end.timestamp())


def _fmt_date(s: str) -> str:
    """ISO/date string -> '5 Jun 2026' for the human-facing brief."""
    if not s:
        return ""
    try:
        d = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        return f"{d.day} {d.strftime('%b %Y')}"
    except (ValueError, AttributeError):
        return s


def _luma_get(path: str, params: dict | None = None) -> dict:
    import requests

    key = os.environ.get("LUMA_API_KEY")
    if not key:
        sys.exit("LUMA_API_KEY is not set (add it to .env). Or run with --mock.")
    r = requests.get(
        f"{LUMA_BASE}{path}",
        headers={"x-luma-api-key": key, "accept": "application/json"},
        params=params or {},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def _luma_paginate(path: str, params: dict) -> list[dict]:
    """Walk Luma cursor pagination, returning all `entries`."""
    out, cursor = [], None
    while True:
        p = dict(params)
        if cursor:
            p["pagination_cursor"] = cursor
        data = _luma_get(path, p)
        out.extend(data.get("entries", data.get("data", [])))
        if not data.get("has_more"):
            break
        cursor = data.get("next_cursor") or data.get("pagination_cursor")
        if not cursor:
            break
    return out


def fetch_luma_sales(start_unix: int, end_unix: int) -> list[dict]:
    """Return one dict per ticket sale within [start, end)."""
    sales: list[dict] = []
    events = _luma_paginate("/calendar/list-events", {})
    for ev in events:
        entry = ev.get("event", ev)
        event_id = entry.get("api_id") or entry.get("event_api_id") or entry.get("id")
        event_name = entry.get("name", "")
        guests = _luma_paginate("/event/get-guests", {"event_api_id": event_id})
        for g in guests:
            guest = g.get("guest", g)
            email = guest.get("email", "")
            name = guest.get("name") or guest.get("user_name") or ""
            detail = _luma_get(
                "/event/get-guest",
                {"event_api_id": event_id, "guest_api_id": guest.get("api_id")},
            ).get("guest", {})
            for order in detail.get("event_ticket_orders", []) or []:
                ts = order.get("created_at") or guest.get("registered_at")
                when = _parse_iso(ts)
                if when is None or not (start_unix <= when.timestamp() < end_unix):
                    continue
                sales.append(
                    {
                        "registered_at": when.isoformat(),
                        "event_id": event_id,
                        "event_name": event_name,
                        "ticket_type": order.get("ticket_type_name", ""),
                        "buyer_name": name,
                        "buyer_email": email,
                        "gross": _money(order.get("amount") or order.get("amount_total")),
                        "currency": (order.get("currency") or "AUD").upper(),
                        "coupon": (order.get("coupon_info") or {}).get("code", ""),
                        "luma_order_id": order.get("api_id", ""),
                    }
                )
    return sales


def _parse_iso(ts):
    if not ts:
        return None
    try:
        return dt.datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except (ValueError, AttributeError):
        return None


def _stripe_get(path: str, params: dict | None = None) -> dict:
    import requests

    key = os.environ.get("STRIPE_API_KEY")
    if not key:
        sys.exit("STRIPE_API_KEY is not set (add it to .env). Or run with --mock.")
    r = requests.get(
        f"{STRIPE_BASE}{path}",
        auth=(key, ""),
        params=params or {},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def _stripe_list(path: str, params: dict) -> list[dict]:
    out, starting_after = [], None
    while True:
        p = dict(params, limit=100)
        if starting_after:
            p["starting_after"] = starting_after
        data = _stripe_get(path, p)
        rows = data.get("data", [])
        out.extend(rows)
        if not data.get("has_more") or not rows:
            break
        starting_after = rows[-1]["id"]
    return out


def fetch_stripe_charges(start_unix: int, end_unix: int) -> list[dict]:
    """Charges grouped by the payout that settled them (payouts in window)."""
    payouts = _stripe_list(
        "/payouts",
        {"arrival_date[gte]": start_unix, "arrival_date[lt]": end_unix},
    )
    charges = []
    for po in payouts:
        txns = _stripe_list(
            "/balance_transactions",
            {"payout": po["id"], "expand[]": "data.source"},
        )
        for t in txns:
            if t.get("type") != "charge":
                continue
            src = t.get("source") or {}
            charges.append(
                {
                    "charge_id": src.get("id") or t["id"],
                    "created": dt.datetime.fromtimestamp(
                        t["created"], dt.timezone.utc
                    ).isoformat(),
                    "gross": _money(t["amount"] / 100),
                    "stripe_fee": _money(t["fee"] / 100),
                    "net": _money(t["net"] / 100),
                    "currency": t["currency"].upper(),
                    "email": (src.get("billing_details") or {}).get("email", ""),
                    "description": src.get("description") or t.get("description") or "",
                    "payout_id": po["id"],
                    "payout_arrival": dt.datetime.fromtimestamp(
                        po["arrival_date"], dt.timezone.utc
                    ).date().isoformat(),
                }
            )
    return charges


def merge(luma_sales: list[dict], stripe_charges: list[dict]) -> list[dict]:
    """One row per Stripe charge (payout-driven), enriched with its Luma sale.

    Payout-driven means every charge in every in-window payout appears, so each
    payout's nets sum to the exact bank deposit. A charge with no Luma match is
    still included, labelled so its event can be picked manually.
    """
    by_order = {s["luma_order_id"]: s for s in luma_sales if s.get("luma_order_id")}
    by_email_amt: dict[tuple, list[dict]] = defaultdict(list)
    for s in luma_sales:
        by_email_amt[(s["buyer_email"].lower(), s["gross"])].append(s)

    rows = []
    for c in stripe_charges:
        s = None
        for oid, sale in by_order.items():
            if oid and oid in (c["description"] or ""):
                s = sale
                break
        if not s:
            cands = by_email_amt.get((c["email"].lower(), c["gross"]), [])
            s = cands[0] if cands else None

        gross = s["gross"] if s else c["gross"]
        luma_fee = _money(gross - c["net"] - c["stripe_fee"]) if s else 0.0
        rows.append(
            {
                "date_bought": s["registered_at"] if s else c["created"],
                "event": s["event_name"] if s else "UNKNOWN (no Luma match)",
                "ticket_type": s["ticket_type"] if s else "",
                "buyer": (s["buyer_email"] or s["buyer_name"]) if s else (c["email"] or "unknown"),
                "gross": gross,
                "luma_fee": luma_fee,
                "stripe_fee": c["stripe_fee"],
                "net": c["net"],
                "currency": c["currency"],
                "stripe_charge_id": c["charge_id"],
                "payout_id": c["payout_id"],
                "payout_arrival": c["payout_arrival"],
            }
        )
    return rows


def payout_summary(rows: list[dict]) -> list[dict]:
    agg: dict[str, dict] = {}
    for r in rows:
        po = r["payout_id"]
        a = agg.setdefault(
            po,
            {
                "payout_id": po,
                "payout_arrival": r["payout_arrival"],
                "tickets": 0,
                "gross": 0.0,
                "luma_fee": 0.0,
                "stripe_fee": 0.0,
                "net": 0.0,
                "events": set(),
            },
        )
        a["tickets"] += 1
        a["gross"] = _money(a["gross"] + r["gross"])
        a["luma_fee"] = _money(a["luma_fee"] + r["luma_fee"])
        a["stripe_fee"] = _money(a["stripe_fee"] + r["stripe_fee"])
        a["net"] = _money(a["net"] + r["net"])
        a["events"].add(r["event"])
    out = []
    for a in agg.values():
        a["events"] = ", ".join(sorted(a["events"]))
        out.append(a)
    return sorted(out, key=lambda x: x["payout_arrival"])


def mock_rows() -> list[dict]:
    raw = [
        ("2026-06-03T09:12:00+00:00", "MLAI Workshop — June", "General",     "alice@example.com", 50.00, 1.15, "ch_3Na001", "po_06051", "2026-06-05"),
        ("2026-06-03T14:40:00+00:00", "MLAI Workshop — June", "General",     "bob@example.com",   50.00, 1.15, "ch_3Na002", "po_06051", "2026-06-05"),
        ("2026-06-04T19:05:00+00:00", "MLAI Workshop — June", "VIP",         "carol@example.com", 120.00, 2.34, "ch_3Na003", "po_06051", "2026-06-05"),
        ("2026-06-10T11:00:00+00:00", "MLAI Studio Mixer",    "Entry",       "dan@example.com",   30.00, 0.81, "ch_3Na004", "po_06121", "2026-06-12"),
        ("2026-06-11T16:22:00+00:00", "MLAI Studio Mixer",    "Entry",       "erin@example.com",  30.00, 0.81, "ch_3Na005", "po_06121", "2026-06-12"),
        ("2026-06-11T16:25:00+00:00", "MLAI Studio Mixer",    "Entry+Drink", "frank@example.com", 45.00, 1.07, "ch_3Na006", "po_06121", "2026-06-12"),
    ]
    rows = []
    for d, ev, tt, buyer, gross, sfee, ch, po, arr in raw:
        luma_fee = _money(gross * 0.02)
        net = _money(gross - luma_fee - sfee)
        rows.append(
            {
                "date_bought": d,
                "event": ev,
                "ticket_type": tt,
                "buyer": buyer,
                "gross": gross,
                "luma_fee": luma_fee,
                "stripe_fee": sfee,
                "net": net,
                "currency": "AUD",
                "stripe_charge_id": ch,
                "payout_id": po,
                "payout_arrival": arr,
            }
        )
    return rows


SALES_COLS = [
    ("date_bought", "Date bought (UTC)"),
    ("event", "Event"),
    ("ticket_type", "Ticket type"),
    ("buyer", "Buyer"),
    ("gross", "Gross"),
    ("luma_fee", "Luma fee"),
    ("stripe_fee", "Stripe fee"),
    ("net", "Net"),
    ("currency", "Ccy"),
    ("stripe_charge_id", "Stripe charge"),
    ("payout_id", "Payout"),
    ("payout_arrival", "Payout date"),
]
PAYOUT_COLS = [
    ("payout_id", "Payout"),
    ("payout_arrival", "Arrival date"),
    ("tickets", "# Tickets"),
    ("gross", "Gross"),
    ("luma_fee", "Luma fee"),
    ("stripe_fee", "Stripe fee"),
    ("net", "Net (bank deposit)"),
    ("events", "Events"),
]


def write_xlsx(rows: list[dict], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    money_cols = {"gross", "luma_fee", "stripe_fee", "net"}
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    wb = Workbook()

    def render(ws, cols, data, total_keys):
        ws.append([label for _, label in cols])
        for c in range(1, len(cols) + 1):
            cell = ws.cell(1, c)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center")
        for r in data:
            ws.append([r.get(k, "") for k, _ in cols])
        trow = ws.max_row + 1
        ws.cell(trow, 1, "TOTAL").font = bold
        for idx, (k, _) in enumerate(cols, start=1):
            if k in total_keys:
                col = get_column_letter(idx)
                ws.cell(trow, idx, f"=SUM({col}2:{col}{ws.max_row-1})").font = bold
        for idx, (k, _) in enumerate(cols, start=1):
            col = get_column_letter(idx)
            width = max(len(str(cols[idx-1][1])), *(len(str(r.get(k, ""))) for r in data)) if data else len(cols[idx-1][1])
            ws.column_dimensions[col].width = min(max(width + 3, 11), 40)
            if k in money_cols or k == "net":
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, idx).number_format = '#,##0.00'
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "Sales detail"
    render(ws1, SALES_COLS, rows, {"gross", "luma_fee", "stripe_fee", "net"})

    ws2 = wb.create_sheet("Payout summary")
    render(ws2, PAYOUT_COLS, payout_summary(rows), {"tickets", "gross", "luma_fee", "stripe_fee", "net"})

    ws3 = wb.create_sheet("About")
    notes = [
        "MLAI — Luma → Stripe reconciliation worksheet",
        "",
        "Sales detail : one row per Stripe charge, linked to its Luma sale and payout.",
        "Payout summary: one row per Stripe payout = the lump sum that lands in your bank.",
        "",
        "Reconcile in Xero against the 'Net (bank deposit)' column on Payout summary —",
        "each payout matches one received line in your bank feed.",
        "",
        "See the *_cowork_brief.md alongside this file for the per-payout Xero",
        "'Create' fields (Who / What / Why / Event Name / Project Name / Tax Rate).",
        "",
        "UNKNOWN (no Luma match) = a Stripe charge with no matching Luma sale found",
        "(check email/amount, refunds, or pick the event manually).",
    ]
    for n in notes:
        ws3.append([n])
    ws3.column_dimensions["A"].width = 80

    wb.save(path)


def write_brief(rows: list[dict], path: str, event_map: dict | None = None) -> None:
    """Write the Cowork-facing markdown brief: one section per Stripe payout."""
    event_map = event_map or {}
    payouts: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        payouts[r["payout_id"]].append(r)
    order = sorted(payouts, key=lambda p: payouts[p][0]["payout_arrival"])

    L: list[str] = []
    L.append("# Stripe payout reconciliation brief")
    L.append("")
    L.append("**For Claude Cowork.** Reconcile each Stripe deposit below in the Xero")
    L.append("bank feed. **Only these Stripe/Luma deposits — leave every other bank")
    L.append("line (transfers, PayID, etc.) untouched.**")
    L.append("")
    L.append("For each: open the bank line → **Create** tab → fill Who / Why, then")
    L.append("**Add details** to split by event. Each split line sets **What** (account),")
    L.append("**Event Name**, **Project Name**, **Amount**, **Tax Rate**. The split total")
    L.append("must equal the bank line to the cent. **A human clicks OK to confirm.**")
    L.append("")
    L.append("> Tax Rate is left as the income account's default — confirm it in Xero.")
    L.append("")

    # Summary table.
    L.append("## Deposits to reconcile")
    L.append("")
    L.append("| Payout | Arrived | Bank deposit | Events | Tickets |")
    L.append("|---|---|---|---|---|")
    for po in order:
        prs = payouts[po]
        ccy = prs[0]["currency"]
        net = _money(sum(r["net"] for r in prs))
        evs = ", ".join(sorted({r["event"] for r in prs}))
        L.append(f"| `{po}` | {_fmt_date(prs[0]['payout_arrival'])} | {ccy} {net:,.2f} | {evs} | {len(prs)} |")
    n_unknown = sum(1 for r in rows if r["event"].startswith("UNKNOWN"))
    if n_unknown:
        L.append("")
        L.append(f"> ⚠ {n_unknown} charge(s) have no Luma match — event must be picked manually (marked below).")
    L.append("")
    L.append("---")
    L.append("")

    for po in order:
        prs = payouts[po]
        ccy = prs[0]["currency"]
        arrival = _fmt_date(prs[0]["payout_arrival"])
        net = _money(sum(r["net"] for r in prs))
        sfee = _money(sum(r["stripe_fee"] for r in prs))
        lfee = _money(sum(r["luma_fee"] for r in prs))
        by_event: dict[str, float] = defaultdict(float)
        for r in prs:
            by_event[r["event"]] = _money(by_event[r["event"]] + r["gross"])
        who = prs[0]["buyer"] if len(prs) == 1 else DEFAULT_PAYER
        events_list = ", ".join(sorted(by_event))

        L.append(f"## {po} — {ccy} {net:,.2f} received {arrival}")
        L.append("")
        L.append(f"**Match the bank line:** Received **{ccy} {net:,.2f}** on/around **{arrival}** (payer: Stripe).")
        L.append("")
        L.append(f"- **Who:** {who}")
        L.append(f"- **Why:** Luma tickets — {events_list} — {len(prs)} ticket(s) — payout {po}")
        L.append("")
        L.append("**Create → Add details (split lines):**")
        L.append("")
        L.append("| What (account) | Event Name | Project Name | Amount | Tax Rate |")
        L.append("|---|---|---|---|---|")
        for event_name in sorted(by_event):
            xero_ev = event_map.get(event_name, "")
            ev_cell = xero_ev if xero_ev else f"⚠ pick — Luma: {event_name}"
            L.append(f"| {TICKET_INCOME_ACCOUNT} | {ev_cell} | (set if used) | {by_event[event_name]:,.2f} | {TAX_RATE_LABEL} |")
        if sfee:
            L.append(f"| {STRIPE_FEE_ACCOUNT} | — | — | -{sfee:,.2f} | {TAX_RATE_LABEL} |")
        if lfee:
            L.append(f"| {LUMA_FEE_ACCOUNT} | — | — | -{lfee:,.2f} | {TAX_RATE_LABEL} |")
        L.append(f"| **TOTAL — must equal the bank line** | | | **{net:,.2f}** | |")
        L.append("")
        L.append("<details><summary>Buyers in this payout (audit — not entered per line)</summary>")
        L.append("")
        for r in sorted(prs, key=lambda x: x["date_bought"]):
            tt = f" / {r['ticket_type']}" if r["ticket_type"] else ""
            L.append(f"- {r['buyer']} — {r['event']}{tt} — {ccy} {r['gross']:,.2f} — {_fmt_date(r['date_bought'])}")
        L.append("")
        L.append("</details>")
        L.append("")
        L.append("> Cowork fills the form; **a human clicks OK to confirm the reconcile.**")
        L.append("")

    with open(path, "w") as f:
        f.write("\n".join(L))


def main() -> None:
    ap = argparse.ArgumentParser(description="Luma + Stripe -> Xero reconciliation pack")
    ap.add_argument("--mock", action="store_true", help="use built-in sample data (no API keys)")
    ap.add_argument("--days", type=int, default=30, help="rolling window size in days (default 30)")
    ap.add_argument("--since", help="window start YYYY-MM-DD (overrides --days)")
    ap.add_argument("--until", help="window end YYYY-MM-DD")
    ap.add_argument("--month", help="single calendar month YYYY-MM (overrides --days/--since)")
    ap.add_argument("--out", default="reconciliation.xlsx", help="output .xlsx path")
    ap.add_argument("--brief", help="output .md brief path (default: <out>_cowork_brief.md)")
    args = ap.parse_args()

    event_map = {"MLAI Workshop — June": "AI Engineer"} if args.mock else _load_event_map()

    if args.mock:
        rows = mock_rows()
    else:
        _load_dotenv()
        start, end = resolve_window(args)
        luma_start = start - LUMA_LOOKBACK_DAYS * 86400
        win = f"{dt.datetime.fromtimestamp(start, dt.timezone.utc).date()} .. {dt.datetime.fromtimestamp(end, dt.timezone.utc).date()}"
        print(f"Window (payout arrival): {win}")
        print("Pulling Stripe payouts/charges ...")
        stripe = fetch_stripe_charges(start, end)
        print(f"  {len(stripe)} Stripe charges across {len({c['payout_id'] for c in stripe})} payouts")
        print("Pulling Luma sales (with lookback) ...")
        luma = fetch_luma_sales(luma_start, end)
        print(f"  {len(luma)} Luma ticket sales")
        rows = merge(luma, stripe)

    write_xlsx(rows, args.out)
    brief_path = args.brief or (os.path.splitext(args.out)[0] + "_cowork_brief.md")
    write_brief(rows, brief_path, event_map)

    n_un = sum(1 for r in rows if r["event"].startswith("UNKNOWN"))
    print(f"Wrote {args.out} and {brief_path}: "
          f"{len(rows)} charges, {len(payout_summary(rows))} payouts"
          + (f", {n_un} with no Luma match" if n_un else ""))


if __name__ == "__main__":
    main()
